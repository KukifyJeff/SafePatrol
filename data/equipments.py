import openpyxl

# === 输入输出文件路径 ===
input_path = "checkitems2.xlsx"
output_path = "equipments2.xlsx"

# === 加载工作簿 ===
wb_in = openpyxl.load_workbook(input_path)
ws_in = wb_in.active

# === 创建输出工作簿 ===
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Unique Equipments"

# 写表头
ws_out.append(["Equipment ID", "Equipment Name", "Point ID"])

# === 读取输入数据并去重 ===
seen = set()
for i, row in enumerate(ws_in.iter_rows(min_row=2), start=2):  # 跳过表头
    equip_id = ws_in[f"C{i}"].value  # 设备ID
    equip_name = ws_in[f"B{i}"].value  # 设备名
    point_id = ws_in[f"A{i}"].value  # 点位ID

    if not equip_id or equip_id in seen:
        continue

    seen.add(equip_id)
    ws_out.append([equip_id, equip_name, point_id])

# === 保存输出 ===
wb_out.save(output_path)
print(f"✅ 处理完成！文件已保存为: {output_path}")
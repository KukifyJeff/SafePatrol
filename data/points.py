import openpyxl
import re

# === 输入输出文件路径 ===
input_path = "equipments2.xlsx"
output_path = "points2.xlsx"

# === 加载工作簿 ===
wb_in = openpyxl.load_workbook(input_path)
ws_in = wb_in.active

# === 创建输出工作簿 ===
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Unique Points"

# 写表头
ws_out.append(["Point ID", "Route ID"])

# === 提取并去重 ===
seen = set()

for i, row in enumerate(ws_in.iter_rows(min_row=2), start=2):  # 跳过表头
    point_id = ws_in[f"C{i}"].value  # 从C列提取 Point ID
    if not point_id or point_id in seen:
        continue

    seen.add(point_id)

    # 去掉末尾的 "-P0x"（如 QH1-O-P01 -> QH1-O）
    route_id = re.sub(r"-P\d{2}$", "", point_id.strip())

    ws_out.append([point_id, route_id])

# === 保存输出 ===
wb_out.save(output_path)
print(f"✅ 处理完成！文件已保存为: {output_path}")
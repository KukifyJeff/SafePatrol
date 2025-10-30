import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

INPUT = "data.xlsx"
OUTPUT = "checkitems.xlsx"

# You can change this if your route prefix is different
PREFIX = "QH2-O"  # e.g. "QH1-O"

# Columns
COL_POINT_ID = "A"
COL_EQUIP_NAME = "B"
COL_EQUIP_ID = "C"
COL_ITEM_NAME = "D"
COL_ITEM_ID = "E"

# Load workbook
wb = openpyxl.load_workbook(INPUT)
ws = wb.active

# Trackers
point_counter = 0
current_point_id = None

# Per-point trackers
equip_counter = 0
device_name_to_eid = {}  # maps device name -> equipment id (e.g. "PumpA" -> "QH1-O-P01E01")
last_device_name = None

# Per-equipment item counters
item_counter_by_eid = {}  # maps equipment id -> next item index (int)

# Iterate rows starting at row 2 (assuming row1 is header)
for row_idx in range(2, ws.max_row + 1):
    point_cell = ws[f"{COL_POINT_ID}{row_idx}"]
    equip_cell = ws[f"{COL_EQUIP_NAME}{row_idx}"]
    item_cell = ws[f"{COL_ITEM_NAME}{row_idx}"]

    # Normalize string values (strip) or None
    point_val = (str(point_cell.value).strip() if point_cell.value is not None else "")

    equip_val = (str(equip_cell.value).strip() if equip_cell.value is not None else "")
    item_val = (str(item_cell.value).strip() if item_cell.value is not None else "")

    # Detect blank row (all three primary columns empty)
    if (not point_val) and (not equip_val) and (not item_val):
        # reset point context so next non-empty row creates a new point id
        current_point_id = None
        last_device_name = None
        device_name_to_eid.clear()
        item_counter_by_eid.clear()
        equip_counter = 0
        continue

    # If current_point_id is None -> start a new point
    if current_point_id is None:
        point_counter += 1
        current_point_id = f"{PREFIX}-P{point_counter:02d}"
        # write point id into column A for this row (and subsequent rows belonging to the same point)
        ws[f"{COL_POINT_ID}{row_idx}"].value = current_point_id
        # reset per-point trackers
        equip_counter = 0
        device_name_to_eid = {}
        last_device_name = None
        item_counter_by_eid = {}
    else:
        # if point column is empty but we have current_point_id, fill it
        if not point_val:
            ws[f"{COL_POINT_ID}{row_idx}"].value = current_point_id
        else:
            # If the sheet already contained a point value that differs from current_point_id, treat as new point
            # (rare if we assume A is empty initially, but handle just in case)
            if point_val != current_point_id:
                point_counter += 1
                current_point_id = f"{PREFIX}-P{point_counter:02d}"
                ws[f"{COL_POINT_ID}{row_idx}"].value = current_point_id
                equip_counter = 0
                device_name_to_eid = {}
                last_device_name = None
                item_counter_by_eid = {}

    # Handle equipment assignment
    if equip_val:
        # If same as last_device_name, reuse the same equipment id (do not increment)
        if last_device_name and equip_val == last_device_name:
            # reuse
            eid = device_name_to_eid.get(equip_val)
            if eid is None:
                # unexpected — if mapping missing, create new
                equip_counter += 1
                eid = f"{current_point_id}E{equip_counter:02d}"
                device_name_to_eid[equip_val] = eid
        else:
            # different device name; if we've seen this device name earlier in this point reuse its id,
            # otherwise allocate a new equipment number
            if equip_val in device_name_to_eid:
                eid = device_name_to_eid[equip_val]
            else:
                equip_counter += 1
                eid = f"{current_point_id}E{equip_counter:02d}"
                device_name_to_eid[equip_val] = eid
        last_device_name = equip_val
        ws[f"{COL_EQUIP_ID}{row_idx}"].value = eid
        # ensure item counter initialized
        if eid not in item_counter_by_eid:
            item_counter_by_eid[eid] = 0
    else:
        # No equipment name on this row: if there was a last device, inherit its eid
        if last_device_name:
            eid = device_name_to_eid.get(last_device_name)
            if eid:
                ws[f"{COL_EQUIP_ID}{row_idx}"].value = eid
            else:
                # nothing to inherit; leave blank
                eid = None
        else:
            eid = None

    # Handle item id
    if item_val:
        if not eid:
            # No equipment id available, try to create one by incrementing (rare)
            equip_counter += 1
            eid = f"{current_point_id}E{equip_counter:02d}"
            ws[f"{COL_EQUIP_ID}{row_idx}"].value = eid
            device_name_to_eid[eid] = eid
            item_counter_by_eid[eid] = 0

        # assign next item number for this equipment
        item_counter_by_eid.setdefault(eid, 0)
        item_counter_by_eid[eid] += 1
        item_idx = item_counter_by_eid[eid]
        iid = f"{eid}I{item_idx:02d}"
        ws[f"{COL_ITEM_ID}{row_idx}"].value = iid

# save
wb.save(OUTPUT)
print("Saved:", OUTPUT)